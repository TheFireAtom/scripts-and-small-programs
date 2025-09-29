from openpyxl import Workbook, load_workbook

workbook = load_workbook("/home/thefireatom/Downloads/IKTST_2_k_vesna_23_24.xlsx")
ws = workbook.active

day_col = "FY"
para_col = "FZ"
subject_col = "GI"
week_col = "GC"
subject_type_col = "GJ"

print("")
print("Расписание группы БПБО-02-22: ", "\n")
print("Нечётная неделя (I): ", "\n")

for subject, day, para, week, subject_type in zip(ws[subject_col], ws[day_col], ws[para_col], ws[week_col], ws[subject_type_col]):

    if (day.value == "ПОНЕДЕЛЬНИК"):
        print(day.value + ":", "\n")

    elif (day.value == "ВТОРНИК"):
        print(f"\n{day.value}:\n")

    elif (day.value == "СРЕДА"):
        print(f"\n{day.value}:\n")

    elif (day.value == "ЧЕТВЕРГ"):
        print(f"\n{day.value}:\n")

    elif (day.value == "ПЯТНИЦА"):
        print(f"\n{day.value}:\n")

    elif (day.value == "СУББОТА"):
        print(f"\n{day.value}:\n")
        
    if (para.value == "1"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "2"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "3"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "4"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "5"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "6"):
        if (week.value == "I"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "7"):
        if (week.value == "I"):
            print(f"{para.value}. \t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")

print()

print("Чётная неделя (II): ", "\n")

for subject, day, para, week, subject_type in zip(ws[subject_col], ws[day_col], ws[para_col], ws[week_col], ws[subject_type_col]):

    if (day.value == "ПОНЕДЕЛЬНИК"):
        print(day.value + ":", "\n")

    elif (day.value == "ВТОРНИК"):
        print(f"\n{day.value}:\n")

    elif (day.value == "СРЕДА"):
        print(f"\n{day.value}:\n")

    elif (day.value == "ЧЕТВЕРГ"):
        print(f"\n{day.value}:\n")

    elif (day.value == "ПЯТНИЦА"):
        print(f"\n{day.value}:\n")

    elif (day.value == "СУББОТА"):
        print(f"\n{day.value}:\n")
        
    if (para.value == "1"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "2"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "3"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "4"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "5"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "6"):
        if (week.value == "II"):
            print(f"{para.value}.\t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
    if (para.value == "7"):
        if (week.value == "II"):
            print(f"{para.value}. \t{subject.value} {subject_type.value}")
        else:
            print(para.value + ".")
